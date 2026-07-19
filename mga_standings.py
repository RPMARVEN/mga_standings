"""
MGA Ryder Cup Points Standings Generator
Reads Results.xlsx + Points table, calculates per-player standings,
outputs HTML and PDF.
"""

import math
import openpyxl
import os
import json
import re
import urllib.request
from collections import defaultdict
from datetime import datetime

# ── CONFIG ──────────────────────────────────────────────────────────────────
RESULTS_FILE = os.path.join(os.path.dirname(__file__), "Results.xlsx")
OUTPUT_DIR   = os.path.dirname(__file__)

# Course location - Eldorado Country Club, McKinney TX
COURSE_LAT = 33.1972
COURSE_LON = -96.6397
COURSE_NAME = "Eldorado Country Club"

# Tournament definitions:
# (display_name, sheet_name, event_type, team_size, places_paid, has_flights, date_str, participation_pts, season)
TOURNAMENTS = [
    # ── 2025-26 Season ──
    ("Member-Member '25",      "Member-Member 2025", "2man_mm",  2,  3,  False, "2025-11-01", 25,  "2025-26"),
    ("2-Man Scramble",         "2-Man Scramble",   "2man",       2,  4,  True,  "2025-12-06", 10,  "2025-26"),
    ("ABCD Scramble",          "ABCD Scramble",    "4man",       4,  5,  False, "2026-01-10", 10,  "2025-26"),
    ("2-Man Shamble",          "2-Man Shamble",    "2man",       2,  4,  True,  "2026-02-21", 10,  "2025-26"),
    ("Lonely Guy - Qualifier", None,               "lonely_guy", 1,  0,  False, "2026-03-07", 0,  "2025-26"),
    ("Quota",                  "Quota",            "individual", 1,  3,  True,  "2026-04-11", 10,  "2025-26"),
    ("The Gerald",             "The Gerald",       "individual", 1,  3,  True,  "2026-05-02", 10,  "2025-26"),
    ("Lonely Guy",             None,               "lonely_guy_mp", 1,  0,  False, "2026-07-12", 50,  "2025-26"),
    ("2 Man Match Play",       None,               "2man_mp", 2,  0,  False, "2026-05-11", 25,  "2025-26"),
    ("Member-Member '26",      "Member-Member 2026", "2man_mm",  2,  3,  True,  "2026-06-06", 25,  "2025-26"),
    ("Presidents Cup",         "Presidents Cup",   "presidents", 1,  5,  True,  "2026-07-11", 50,  "2025-26"),
    ("6-6-6",                  None,               "tbd",        0,  0,  False, "2026-08-08", 10,  "2025-26"),
    ("Eldo Cup",               None,               "eldo_cup",   0,  0,  False, "2026-09-18", 20,  "2025-26"),
    # ── 2026-27 Season ──
    ("Member-Member-Member",   None,               "3man",       3,  4,  False, "2026-10-17", 10,  "2026-27"),
    ("2-Man RWB",              None,               "2man",       2,  4,  False, "2026-11-14", 10,  "2026-27"),
    ("Mulligan",               None,               "individual", 1,  4,  False, "2026-12-12", 10,  "2026-27"),
]

# Schedule annotations
SCHEDULE_NOTES = {
    "Eldo Cup": "Qualification / Captain's picks",
    "6-6-6": "Last qualifying event",
}

MULTI_DAY_END = {
    "Member-Member '25": "2025-11-02",
    "Member-Member '26": "2026-06-07",
    "Presidents Cup": "2026-07-12",
    "Eldo Cup": "2026-09-20",
}

CANCELLED_EVENTS = {"Lonely Guy - Qualifier"}

# ── Match-play bracket results ───────────────────────────────────────────────
# Entered manually from the GolfGenius bracket JPGs (sources/Lonely_Guy_*.jpg,
# sources/2_Man_Match_Play_*.jpg). Each round win is scored as it completes;
# remaining rounds are appended to round_winners as they are played.
#
# Per-player scoring (points_structure/RYAN PARKS RYDER CUP RECOMMENDATIONS.xlsx ladder):
#   - Participation (set per-tournament in TOURNAMENTS) to EVERY entrant.
#   - round_pts[r] added for each round r the player/team WON.
#   Each round win is a flat step on the ladder (the 2x only applies at the
#   Final/Champion, which are not yet played):
#   Lonely Guy: 64-player field, +50/round. R1 Field/32, R2 Field/16,
#               R3 Field/8 (Round of 16, Jun 14), R4 Field/4 (Quarterfinals, Jul 12).
#   2 Man MP:   32-team field, +25/round. R1 Field/16, R2 Field/8,
#               R3 Field/4 (Quarterfinals), R4 Field/2 (Semifinals -> finalists).
# Names are normalized to the canonical "Last, First" roster used across events.
MATCH_PLAY_TYPES = {"lonely_guy_mp", "2man_mp"}

MATCH_PLAY_RESULTS = {
    "Lonely Guy": {
        "round_pts": [50, 50, 50, 50],
        "entrants": [
            # 32 R1 matches (winner listed first in each pair)
            "Parks, Ryan", "Trevino, Bradley",
            "Harnett, William", "Dorrance, Will",
            "Wojnas, Jay", "Lanford, Mike",
            "Ludwig, George", "Collins, Don",
            "Harper, Ron", "Byrns, Ray",
            "Jones, Brent", "Phillips, Jason",
            "Russell, Jr", "Kelley, Mike",
            "Cottam, Ryan", "Tryggestad, Devon",
            "Quarles, Aaron", "Badgley, Rick",
            "Bolen, Trey", "Peirson, Matt",
            "Fry, Brett", "Rush, Jim",
            "Wells, Daniel", "McCall, Cam",
            "Parks, Mike", "Nelson, Brandon",
            "Coffin, Greg", "Liepins, Eric",
            "Donovan, John", "Girot, Robert",
            "Bruegel, Michael", "Novich, Darren",
            "Newbrough, Stu", "Hawley, Clay",
            "Spencer, Steve", "Novak, Dite",
            "Finnin, Michael", "Cross, Jeremy",
            "Lowe, Kevin", "Dunsworth, Michael",
            "Harless, Garrett", "Mendoza, Daniel",
            "Girot, Geoffrey", "Tredinnick, Rick",
            "Brown, Chris", "Ampaipitakwong, Pete",
            "Finn, Kelly", "Agnew, Mark",
            "Dowling, Patrick", "Weiss, Brian",
            "Miles, Ryan", "Recker, Thomas",
            "Anders, Ken", "Kiernan, John",
            "Bennett, Dustan", "Lavalette, John",
            "Morrone, Jeff", "Flanders, Andrew",
            "Locke, Tom", "Thomas, Bobby",
            "Butler, Terry", "Moore, Joseph",
            "Michelsen, John", "Larkin, Tripp",
        ],
        "round_winners": [
            # R1 winners (32)
            [
                "Parks, Ryan", "Harnett, William", "Wojnas, Jay", "Ludwig, George",
                "Harper, Ron", "Jones, Brent", "Russell, Jr", "Cottam, Ryan",
                "Quarles, Aaron", "Bolen, Trey", "Fry, Brett", "Wells, Daniel",
                "Parks, Mike", "Coffin, Greg", "Donovan, John", "Bruegel, Michael",
                "Newbrough, Stu", "Spencer, Steve", "Finnin, Michael", "Lowe, Kevin",
                "Harless, Garrett", "Girot, Geoffrey", "Brown, Chris", "Finn, Kelly",
                "Dowling, Patrick", "Miles, Ryan", "Anders, Ken", "Bennett, Dustan",
                "Morrone, Jeff", "Locke, Tom", "Butler, Terry", "Michelsen, John",
            ],
            # R2 winners (16)
            [
                "Harnett, William", "Ludwig, George", "Jones, Brent", "Cottam, Ryan",
                "Quarles, Aaron", "Wells, Daniel", "Parks, Mike", "Bruegel, Michael",
                "Spencer, Steve", "Lowe, Kevin", "Harless, Garrett", "Brown, Chris",
                "Miles, Ryan", "Anders, Ken", "Morrone, Jeff", "Butler, Terry",
            ],
            # R3 winners (8) - Round of 16, played Jun 14
            [
                "Harnett, William", "Jones, Brent", "Quarles, Aaron", "Bruegel, Michael",
                "Lowe, Kevin", "Brown, Chris", "Anders, Ken", "Butler, Terry",
            ],
            # R4 winners (4) - Quarterfinals, played Jul 12 (semifinalists)
            [
                "Harnett, William", "Bruegel, Michael", "Brown, Chris", "Butler, Terry",
            ],
        ],
    },
    "2 Man Match Play": {
        "round_pts": [25, 25, 25, 25],
        "entrants": [
            # 32 R1 matches (winner listed first in each pair)
            "Harless, Garrett + Harnett, William", "Girot, Geoffrey + Girot, Robert",
            "Brown, Chris + Finnin, Michael", "Tryggestad, Devon + Cross, Jeremy",
            "Ludwig, George + Novak, Dite", "Agnew, Mark + Morrone, Jeff",
            "Trevino, Bradley + Quarles, Aaron", "Alt, Kyle + Switser, Chris",
            "Dunsworth, Michael + Wells, Daniel", "Dowling, Patrick + Ampaipitakwong, Pete",
            "Wojnas, Jay + Locke, Tom", "Terranova, Adrian + Newbrough, Stu",
            "Dorrance, Will + Springer, Chris", "Rush, Jim + Avila, Tom",
            "Lang, Stuart + Adler, John", "Bausch, Marty + Thomas, Bobby",
            "Anders, Ken + Kiernan, John", "Parks, Mike + Cottam, Ryan",
            "Byrns, Ray + Desing, Patrick", "Bolen, Trey + Mueller, Shawn",
            "Coffin, Greg + Hawley, Clay", "Jones, Brent + Lowe, Kevin",
            "Swartz, Michael + Spencer, Steve", "Butler, Terry + Kelley, Mike",
            "Jones, Steve + Meredith, Mike", "Phillips, Jason + Larkin, Tripp",
            "Bruegel, Michael + Strickland, Joshua", "Nelson, Brandon + Badgley, Rick",
            "Recker, Thomas + Russell, Jr", "Wood, Peter + Lanford, Mike",
            "Collins, Don + Lavalette, John", "Novich, Darren + Reyes, Marvin",
        ],
        "round_winners": [
            # R1 winning teams (16)
            [
                "Harless, Garrett + Harnett, William", "Alt, Kyle + Switser, Chris",
                "Dunsworth, Michael + Wells, Daniel", "Dorrance, Will + Springer, Chris",
                "Rush, Jim + Avila, Tom", "Anders, Ken + Kiernan, John",
                "Parks, Mike + Cottam, Ryan", "Coffin, Greg + Hawley, Clay",
                "Jones, Brent + Lowe, Kevin", "Swartz, Michael + Spencer, Steve",
                "Phillips, Jason + Larkin, Tripp", "Bruegel, Michael + Strickland, Joshua",
                "Nelson, Brandon + Badgley, Rick", "Recker, Thomas + Russell, Jr",
                "Wood, Peter + Lanford, Mike", "Collins, Don + Lavalette, John",
            ],
            # R2 winning teams (8)
            [
                "Dunsworth, Michael + Wells, Daniel", "Dorrance, Will + Springer, Chris",
                "Rush, Jim + Avila, Tom", "Anders, Ken + Kiernan, John",
                "Jones, Brent + Lowe, Kevin", "Swartz, Michael + Spencer, Steve",
                "Nelson, Brandon + Badgley, Rick", "Collins, Don + Lavalette, John",
            ],
            # R3 winning teams (4) - Quarterfinals
            [
                "Anders, Ken + Kiernan, John", "Dorrance, Will + Springer, Chris",
                "Collins, Don + Lavalette, John", "Jones, Brent + Lowe, Kevin",
            ],
            # R4 winning teams (2) - Semifinals (finalists)
            [
                "Dorrance, Will + Springer, Chris", "Jones, Brent + Lowe, Kevin",
            ],
        ],
    },
}


def event_has_data(t):
    """True if a tournament has results to score: a real sheet OR match-play bracket data."""
    return t[1] is not None or (t[2] in MATCH_PLAY_TYPES and t[0] in MATCH_PLAY_RESULTS)


# Schedule display for the ongoing match-play brackets: the completed rounds are
# shown under Completed; the next round to be played is shown under Ongoing.
# Update done_* and next_round / next_date as later rounds finish.
MATCH_PLAY_ROUNDS = {
    "Lonely Guy": {
        "label": "Lonely Guy",
        "team_size": 1,
        "done_dates": "Apr 30 - Jul 12",
        "done_label": "Rounds 1-4",
        "done_best": "250",   # max per-player banked through round 4 (50 part + 4x50)
        "part": "50",
        "next_round": "Semifinals",   # 4 players remain -> 2
        "next_date": "Aug 2",
    },
    "2 Man Match Play": {
        "label": "2 Man Match Play",
        "team_size": 2,
        "done_dates": "Mar 21 - Jul 12",
        "done_label": "Rounds 1-4",
        "done_best": "125",   # max per-player banked through round 4 (25 part + 4x25)
        "part": "25",
        "next_round": "Final",   # 2 finalists remain -> champion
        "next_date": "TBD",
    },
}

# Points tables (per-player values, already divided)
# From "points_structure/RYAN PARKS RYDER CUP RECOMMENDATIONS.xlsx"
POINTS_TABLE = {
    # (event_type, places_paid): {place: per_player_pts}
    # Individual events
    ("individual", 3): {1: 100, 2: 80, 3: 60},
    ("individual", 4): {1: 100, 2: 80, 3: 60, 4: 40},
    ("individual", 5): {1: 100, 2: 80, 3: 60, 4: 40, 5: 20},
    # 2-man events
    ("2man", 3): {1: 100, 2: 80, 3: 60},
    ("2man", 4): {1: 100, 2: 80, 3: 60, 4: 40},
    ("2man", 5): {1: 100, 2: 80, 3: 60, 4: 40, 5: 10},
    # 3-man events
    ("3man", 3): {1: 100, 2: 80, 3: 60},
    ("3man", 4): {1: 100, 2: 80, 3: 60, 4: 40},
    ("3man", 5): {1: 40, 2: 30, 3: 25, 4: 40, 5: 20},
    # 4-man events
    ("4man", 4): {1: 100, 2: 80, 3: 60, 4: 40},
    ("4man", 5): {1: 100, 2: 80, 3: 60, 4: 40, 5: 20},
    ("4man", 6): {1: 100, 2: 80, 3: 60, 4: 40, 5: 20, 6: 10},
    # Member/Member (special - higher stakes, per-player)
    ("2man_mm", 3): {1: 125, 2: 75, 3: 50},
    # Presidents Cup (individual, major)
    ("presidents", 3): {1: 300, 2: 240, 3: 180},
    ("presidents", 4): {1: 300, 2: 240, 3: 180, 4: 120},
    ("presidents", 5): {1: 300, 2: 240, 3: 180, 4: 120, 5: 60},
}

PARTICIPATION_PTS = 10  # default per player per event (overridden per tournament)

# Available places_paid options per event type (from points table keys)
PLACES_OPTIONS = {
    "individual": [3, 4, 5],
    "2man":       [3, 4, 5],
    "3man":       [3, 4, 5],
    "4man":       [4, 5, 6],
}

def calc_places_paid(event_type, total_teams, num_flights=1):
    """Pick places_paid targeting 1/3 of the field, minimum 3.
    For flighted events, uses teams per flight (equal payout per flight).
    Returns a single places_paid value applied uniformly across all flights."""
    options = PLACES_OPTIONS.get(event_type)
    if not options:
        return None  # special events (2man_mm, presidents, etc.) set manually
    teams_per_flight = total_teams / max(num_flights, 1)
    target = teams_per_flight / 3
    best = min(options, key=lambda x: abs(x - target))
    return max(best, 3)  # minimum 3 places paid

# WMO weather code to description
WMO_CODES = {
    0: "Clear", 1: "Mostly Clear", 2: "Partly Cloudy", 3: "Overcast",
    45: "Foggy", 48: "Foggy", 51: "Light Drizzle", 53: "Drizzle",
    55: "Heavy Drizzle", 61: "Light Rain", 63: "Rain", 65: "Heavy Rain",
    71: "Light Snow", 73: "Snow", 75: "Heavy Snow", 80: "Showers",
    81: "Heavy Showers", 82: "Violent Showers", 95: "Thunderstorm",
}


def fetch_weather(date_str):
    """Fetch weather for a tournament date (morning to early afternoon). Returns dict or None."""
    if not date_str:
        return None
    try:
        url = (f"https://archive-api.open-meteo.com/v1/archive?"
               f"latitude={COURSE_LAT}&longitude={COURSE_LON}"
               f"&start_date={date_str}&end_date={date_str}"
               f"&hourly=temperature_2m,weathercode,windspeed_10m"
               f"&temperature_unit=fahrenheit&timezone=America/Chicago")
        resp = urllib.request.urlopen(url, timeout=10)
        data = json.loads(resp.read())
        temps = data["hourly"]["temperature_2m"]
        winds = data["hourly"]["windspeed_10m"]
        codes = data["hourly"]["weathercode"]
        times = data["hourly"]["time"]
        # Tournament hours: 7am to 2pm
        idx = [i for i, t in enumerate(times) if 7 <= int(t.split("T")[1].split(":")[0]) <= 14]
        if not idx:
            return None
        mt = [temps[i] for i in idx if temps[i] is not None]
        mw = [winds[i] for i in idx if winds[i] is not None]
        mc = [codes[i] for i in idx if codes[i] is not None]
        # Most common weather code
        dominant_code = max(set(mc), key=mc.count) if mc else 0
        return {
            "high": round(max(mt)),
            "low": round(min(mt)),
            "wind_avg": round(sum(mw) / len(mw)),
            "condition": WMO_CODES.get(dominant_code, "Unknown"),
        }
    except Exception:
        return None


def parse_position(pos_val):
    """Parse position like 'T1', 'T2', 1, 2 etc. Returns (numeric_pos, is_tie)."""
    if pos_val is None:
        return None, False
    s = str(pos_val).strip()
    is_tie = s.startswith("T")
    num = int(s.replace("T", ""))
    return num, is_tie


def split_players(players_str):
    """Split 'Last, First + Last, First' into individual player names."""
    parts = players_str.split("+")
    return [p.strip() for p in parts if p.strip()]


def parse_flight_results(rows):
    """Parse rows for a single flight. Returns list of (position_str, [players])."""
    results = []
    for row in rows:
        # row = [link_or_none, pos, players, to_par, thru, total]
        pos_val = row[1]
        players_str = row[2]
        if pos_val is None or players_str is None:
            continue
        # Skip header rows
        if str(pos_val).strip() in ("Pos.", ""):
            continue
        players = split_players(str(players_str))
        results.append((pos_val, players))
    return results


def _get_place_pts(table, places_paid, place):
    """Get points for a place, extending beyond paid places using the established
    increment between consecutive places, flooring at 0."""
    if place in table:
        return table[place]
    if place > places_paid:
        # Determine the increment from the last two paid places
        if places_paid >= 2 and places_paid in table and (places_paid - 1) in table:
            increment = table[places_paid - 1] - table[places_paid]
        elif places_paid >= 1 and places_paid in table:
            increment = 20  # fallback
        else:
            return 0
        last_paid_pts = table.get(places_paid, 0)
        steps_beyond = place - places_paid
        return max(0, last_paid_pts - increment * steps_beyond)
    return 0


def calc_points_for_flight(results, event_type, places_paid):
    """
    Given flight results and point table params, calculate per-player points.
    Tie rules:
    - Tie within paid places: average points across all positions occupied.
      E.g. 4-way tie for 1st with 4 places = (100+80+60+40)/4 = 70 each. Flight done.
    - Tie at last paid place straddling the cutoff: extend phantom places using the
      established increment (floored at 0), then average.
      E.g. 2-way tie for 4th = (40+20)/2 = 30 each.
    Returns dict: {player_name: points}
    """
    table = POINTS_TABLE.get((event_type, places_paid), {})
    player_points = {}

    # results is [(pos_val, [players]), ...]
    i = 0
    while i < len(results):
        pos_val, players = results[i]
        pos, is_tie = parse_position(pos_val)
        if pos is None:
            i += 1
            continue

        # Collect all teams at this tied position
        tied_teams = [(pos_val, players)]
        j = i + 1
        while j < len(results):
            next_pos_val, next_players = results[j]
            next_pos, next_is_tie = parse_position(next_pos_val)
            if next_pos == pos:
                tied_teams.append((next_pos_val, next_players))
                j += 1
            else:
                break

        num_tied = len(tied_teams)
        if pos > places_paid:
            # Entirely outside paid places - no points
            pts = 0
        elif num_tied == 1:
            # No tie - straight payout
            pts = table.get(pos, 0)
        else:
            # Average points across the positions these teams occupy
            # Uses phantom places (extending by established increment, floor 0)
            # when the tie straddles the payout cutoff
            total_pts = sum(_get_place_pts(table, places_paid, pos + offset) for offset in range(num_tied))
            pts = total_pts / num_tied

        for _, team_players in tied_teams:
            for player in team_players:
                player_points[player] = pts

        i = j

    return player_points


def parse_flighted_sheet(ws):
    """Parse a sheet with Flight 1, Flight 2, etc. Returns list of flights, each a list of result rows."""
    flights = []
    current_flight_rows = []
    in_flight = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        # Detect flight header - can be in column A or B depending on sheet format
        is_flight_header = (
            (vals[0] and str(vals[0]).strip().startswith("Flight")) or
            (vals[1] and str(vals[1]).strip().startswith("Flight"))
        )
        if is_flight_header:
            if current_flight_rows:
                flights.append(current_flight_rows)
            current_flight_rows = []
            in_flight = True
            continue
        # Skip sub-headers
        if vals[1] and str(vals[1]).strip() in ("Pos.",):
            continue
        if vals[2] and str(vals[2]).strip() in ("Net",):
            continue
        if in_flight and vals[1] is not None and vals[2] is not None:
            current_flight_rows.append(vals)

    if current_flight_rows:
        flights.append(current_flight_rows)

    return flights


def parse_single_flight_sheet(ws):
    """Parse a sheet with no flights (one big group). Returns list of result rows."""
    results = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        # Skip headers
        if vals[0] and str(vals[0]).strip() in ("Pos.",):
            continue
        if vals[1] and str(vals[1]).strip() in ("Net",):
            continue
        # Data rows have a URL or position in col 0, and a number/Txx in col 1
        pos_val = vals[0] if vals[1] is None else vals[1]
        players_str = vals[1] if vals[2] is None else vals[2]

        # For ABCD sheet: col0=link/pos, col1=pos_number, col2=foursome
        if vals[0] and "golfgenius" in str(vals[0]):
            pos_val = vals[1]
            players_str = vals[2]
        elif vals[0] is not None:
            pos_val = vals[0]
            players_str = vals[1]
        else:
            continue

        if pos_val is None or players_str is None:
            continue
        try:
            parse_position(pos_val)
        except (ValueError, TypeError):
            continue

        players = split_players(str(players_str))
        results.append((pos_val, players))

    return results


def parse_member_member_sheet(ws, event_type, places_paid):
    """
    Parse Member-Member match play sheet.
    Column A has match play brackets with all participants in "Last, First + Last, First" format.
    Columns G & H have the definitive finishing places and player names in "First Last" format.
    Each player's G/H place maps directly to the points table - no tie averaging.
    Returns (all_players set, player_points dict).
    """
    table = POINTS_TABLE.get((event_type, places_paid), {})

    # ── 1. Extract all player names from column A ──
    all_players = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        cell = row[0]
        if cell is None:
            continue
        text = str(cell).strip()
        if text.startswith("Flight") or text.startswith("Pos."):
            continue
        cleaned = re.sub(r'^T?\d+\s+', '', text)
        parts = cleaned.split('+')
        for part in parts:
            m = re.match(r"\s*([A-Za-z\u0027'-]+(?:\s+[A-Za-z\u0027'-]+)*),\s*([A-Za-z\u0027'-]+)", part.strip())
            if m:
                name = f"{m.group(1).strip()}, {m.group(2).strip()}"
                all_players.add(name)

    # ── 2. Build last-name lookup for fuzzy matching ──
    last_name_lookup = {}
    for name in all_players:
        last = name.split(',')[0].strip().lower()
        last_name_lookup.setdefault(last, []).append(name)

    # ── 3. Parse finishing places from columns G & H ──
    player_points = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=8, values_only=True):
        place_val, player_name = row
        if place_val is None or player_name is None:
            continue
        if not isinstance(place_val, (int, float)):
            continue
        place = int(place_val)
        pname = str(player_name).strip()

        # Convert "First Last" to "Last, First"
        name_parts = pname.rsplit(None, 1)
        if len(name_parts) == 2:
            converted = f"{name_parts[1]}, {name_parts[0]}"
        else:
            converted = pname

        # Match to canonical column-A name: exact first, then last-name fallback
        if converted in all_players:
            canonical = converted
        else:
            last_key = name_parts[1].lower() if len(name_parts) == 2 else converted.lower()
            candidates = last_name_lookup.get(last_key, [])
            canonical = candidates[0] if len(candidates) == 1 else converted

        player_points[canonical] = table.get(place, 0)

    return all_players, player_points


def process_all():
    """Main processing: read results, calculate points, return standings."""
    wb = openpyxl.load_workbook(RESULTS_FILE, data_only=True)

    # player -> {tournament_name: placement_points}
    player_data = defaultdict(lambda: defaultdict(float))
    # player -> set of tournament names they participated in
    player_events = defaultdict(set)

    for display_name, sheet_name, event_type, team_size, places_paid, has_flights, _date, _part_pts, _season in TOURNAMENTS:
        # Match-play bracket events (scored from MATCH_PLAY_RESULTS, no sheet)
        if event_type in MATCH_PLAY_TYPES and display_name in MATCH_PLAY_RESULTS:
            mp = MATCH_PLAY_RESULTS[display_name]
            all_players = set()
            for team in mp["entrants"]:
                all_players.update(split_players(team))
            for r, winners in enumerate(mp["round_winners"]):
                pts = mp["round_pts"][r]
                for team in winners:
                    for player in split_players(team):
                        player_data[player][display_name] += pts
            for player in all_players:
                player_data[player][display_name] += 0  # ensure column key exists
                player_events[player].add(display_name)
            continue

        if sheet_name is None:
            continue  # placeholder tournament
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]

        # Detect Member-Member match play format (Results header in col G)
        g1 = ws.cell(1, 7).value
        if g1 and str(g1).strip() == "Results":
            all_participants, pts = parse_member_member_sheet(ws, event_type, places_paid)
            for player in all_participants:
                earned = pts.get(player, 0)
                player_data[player][display_name] += earned
                player_events[player].add(display_name)
            continue

        if has_flights:
            flights = parse_flighted_sheet(ws)
            for flight_rows in flights:
                flight_results = []
                for vals in flight_rows:
                    pos_val = vals[1]
                    players_str = vals[2]
                    if pos_val is None or players_str is None:
                        continue
                    players = split_players(str(players_str))
                    flight_results.append((pos_val, players))

                pts = calc_points_for_flight(flight_results, event_type, places_paid)
                all_players_in_flight = set()
                for _, plist in flight_results:
                    all_players_in_flight.update(plist)
                for player in all_players_in_flight:
                    earned = pts.get(player, 0)
                    player_data[player][display_name] += earned
                    player_events[player].add(display_name)
        else:
            results = parse_single_flight_sheet(ws)
            pts = calc_points_for_flight(results, event_type, places_paid)
            all_players = set()
            for _, plist in results:
                all_players.update(plist)
            for player in all_players:
                earned = pts.get(player, 0)
                player_data[player][display_name] += earned
                player_events[player].add(display_name)

    return player_data, player_events


def build_standings(player_data, player_events):
    """Build sorted standings list."""
    tournament_names = [t[0] for t in TOURNAMENTS if event_has_data(t)]
    part_pts_map = {t[0]: t[7] for t in TOURNAMENTS}

    standings = []
    for player, tourney_pts in player_data.items():
        total_tourney = sum(tourney_pts.values())
        events_set = player_events.get(player, set())
        events = len(events_set)
        participation = sum(part_pts_map.get(e, PARTICIPATION_PTS) for e in events_set)
        standings.append({
            "player": player,
            "tournaments": {t: tourney_pts.get(t, 0) for t in tournament_names},
            "events": events,
            "participation": participation,
            "total": total_tourney + participation,
        })

    # Sort by total descending
    standings.sort(key=lambda x: (-x["total"], x["player"]))

    # Assign ranks with ties
    for i, entry in enumerate(standings):
        if i == 0:
            entry["rank"] = 1
        elif entry["total"] == standings[i - 1]["total"]:
            entry["rank"] = standings[i - 1]["rank"]
        else:
            entry["rank"] = i + 1

    return standings, tournament_names


def _weather_icon_html(condition):
    """Return a self-contained inline SVG weather icon (28x28)."""
    c = condition.lower()
    sz = 28
    svg = f'<svg width="{sz}" height="{sz}" viewBox="0 0 28 28" xmlns="http://www.w3.org/2000/svg">'

    sun = '<circle cx="14" cy="12" r="5" fill="#f59e0b"/>'
    rays = ""
    for angle in range(0, 360, 45):
        rad = math.radians(angle)
        x1, y1 = 14 + 7 * math.cos(rad), 12 + 7 * math.sin(rad)
        x2, y2 = 14 + 9 * math.cos(rad), 12 + 9 * math.sin(rad)
        rays += f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" stroke="#f59e0b" stroke-width="1.5" stroke-linecap="round"/>'

    cloud = '<ellipse cx="12" cy="16" rx="8" ry="4.5" fill="#94a3b8"/><ellipse cx="16" cy="14" rx="6" ry="4" fill="#b0bec5"/><ellipse cx="13" cy="12.5" rx="5" ry="3.5" fill="#cfd8dc"/>'
    cloud_sm = '<ellipse cx="17" cy="17" rx="7" ry="4" fill="#b0bec5"/><ellipse cx="19" cy="15" rx="5" ry="3.5" fill="#cfd8dc"/>'

    rain2 = '<line x1="10" y1="21" x2="9" y2="25" stroke="#3b82f6" stroke-width="1.5" stroke-linecap="round"/><line x1="15" y1="22" x2="14" y2="26" stroke="#3b82f6" stroke-width="1.5" stroke-linecap="round"/>'
    rain3 = rain2 + '<line x1="20" y1="21" x2="19" y2="25" stroke="#3b82f6" stroke-width="1.5" stroke-linecap="round"/>'
    snow = '<circle cx="10" cy="22" r="1.2" fill="#93c5fd"/><circle cx="15" cy="24" r="1.2" fill="#93c5fd"/><circle cx="20" cy="22" r="1.2" fill="#93c5fd"/>'

    if c == "clear":
        svg += sun + rays
    elif c == "mostly clear":
        svg += sun + rays + '<ellipse cx="20" cy="17" rx="5" ry="3" fill="#cfd8dc" opacity="0.6"/>'
    elif c == "partly cloudy":
        svg += '<circle cx="10" cy="10" r="4.5" fill="#f59e0b"/>' + cloud_sm
    elif c == "overcast":
        svg += cloud
    elif c == "foggy":
        for dy in range(0, 4):
            y = 9 + dy * 4
            svg += f'<line x1="5" y1="{y}" x2="23" y2="{y}" stroke="#b0bec5" stroke-width="2" stroke-linecap="round" opacity="0.6"/>'
    elif c in ("light drizzle", "drizzle"):
        svg += cloud + rain2
    elif c in ("heavy drizzle", "light rain", "rain", "showers"):
        svg += cloud + rain3
    elif c in ("heavy rain", "heavy showers", "violent showers"):
        svg += cloud + rain3 + '<line x1="12" y1="23" x2="11" y2="27" stroke="#3b82f6" stroke-width="1.5" stroke-linecap="round"/>'
    elif "snow" in c:
        svg += cloud + snow
    elif c == "thunderstorm":
        svg += cloud + '<polygon points="13,19 16,23 14,23 16,27" fill="#facc15" stroke="#eab308" stroke-width="0.5"/>' + rain2
    else:
        svg += cloud

    svg += '</svg>'
    return svg


def format_rank_delta(delta):
    """Render rank-change cell HTML. delta > 0 = moved up, < 0 = down, None = new entrant."""
    if delta is None:
        return '<td class="rank-delta new">NEW</td>'
    if delta > 0:
        return f'<td class="rank-delta up">&#9650; {delta}</td>'
    if delta < 0:
        return f'<td class="rank-delta down">&#9660; {abs(delta)}</td>'
    return '<td class="rank-delta flat">-</td>'


def build_overview_html(standings, player_data, player_events, rank_changes):
    """Build the overview page: left stats + chart, right top 10."""
    total_players = len(standings)
    current_season = "2025-26"
    season_tournaments = [t for t in TOURNAMENTS if t[8] == current_season]
    completed = [t for t in season_tournaments if t[1] is not None]
    total_scheduled = len(season_tournaments)
    num_completed = len(completed)

    # Field sizes per completed tournament
    field_stats = get_field_stats()
    field_sizes = [field_stats[t[0]]["total_players"] for t in completed if t[0] in field_stats]
    avg_field = round(sum(field_sizes) / len(field_sizes)) if field_sizes else 0

    # Next upcoming tournament (future, unplayed, not cancelled)
    today_str = datetime.now().strftime("%Y-%m-%d")
    next_tourney = None
    for t in TOURNAMENTS:
        if t[1] is None and t[6] and t[6] >= today_str and t[0] not in CANCELLED_EVENTS:
            next_tourney = t
            break

    # Weather for completed tournaments
    print("  Fetching weather data...")
    weather_data = {}
    for t in completed:
        w = fetch_weather(t[6])
        if w:
            weather_data[t[0]] = w

    # Top 15 (hard cutoff at 15, no ties beyond)
    top15_rows = ""
    for entry in standings[:15]:
        rank = entry["rank"]
        rank_display = f"T{rank}" if any(
            e["rank"] == rank and e["player"] != entry["player"] for e in standings
        ) else str(rank)
        total_display = int(entry["total"]) if entry["total"] == int(entry["total"]) else f"{entry['total']:.1f}"
        delta_cell = format_rank_delta(rank_changes.get(entry["player"]))
        top15_rows += f"""
                <tr>
                    <td class="rank">{rank_display}</td>
                    {delta_cell}
                    <td class="player">{entry['player']}</td>
                    <td class="events">{entry['events']}</td>
                    <td class="total">{total_display}</td>
                </tr>"""

    # Stat cards
    stats_html = ""
    stat_items = [
        ("Total Players", str(total_players)),
        ("Season Progress", f"{num_completed} of {total_scheduled}"),
        ("Avg Field Size", str(avg_field)),
    ]
    for label, value in stat_items:
        stats_html += f"""
                <div class="stat-card">
                    <div class="stat-number">{value}</div>
                    <div class="stat-label">{label}</div>
                </div>"""

    # ── Field Size + Conditions (full width, forecast style) ──
    fc_cards = ""
    for i, t in enumerate(completed):
        name = t[0]
        date_str = t[6]
        players = field_stats.get(name, {}).get("total_players", 0)
        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_label = dt.strftime("%b %d")
        else:
            date_label = ""
        w = weather_data.get(t[0])
        weather_icon = ""
        weather_detail = ""
        if w:
            weather_icon = f'<div class="fc-icon">{_weather_icon_html(w["condition"])}</div>'
            weather_detail = f"""<div class="fc-temps">{w['high']}° / {w['low']}°</div>
                <div class="fc-wind">{w['wind_avg']} mph</div>"""

        fc_cards += f"""
            <div class="fc-card">
                <div class="fc-name">{name}</div>
                <div class="fc-date">{date_label}</div>
                <div class="fc-field">{players}</div>
                <div class="fc-field-label">players</div>
                {weather_icon}
                {weather_detail}
            </div>"""

    pc_html = f"""
        <div class="fc-container">
            {fc_cards}
        </div>"""

    # Next Up card
    next_up_html = ""
    if next_tourney:
        nt_name = next_tourney[0]
        nt_team = next_tourney[3]
        nt_type = "Individual" if nt_team == 1 else (f"{nt_team}-Man" if nt_team > 0 else "Special")
        nt_places = next_tourney[4]
        nt_part = next_tourney[7]
        table_key = (next_tourney[2], nt_places)
        pts_table = POINTS_TABLE.get(table_key, {})
        first_pp = pts_table.get(1, 0)
        if first_pp > 0 and nt_team > 1:
            pts_line = f"1st: {int(first_pp * nt_team)} total ({int(first_pp)}/player)"
        elif first_pp > 0:
            pts_line = f"1st: {int(first_pp)} pts"
        else:
            pts_line = ""
        if next_tourney[6]:
            nt_dt = datetime.strptime(next_tourney[6], "%Y-%m-%d")
            nt_date_display = nt_dt.strftime("%a %b %d, %Y")
        else:
            nt_date_display = ""
        next_up_html = f"""
        <div class="next-up">
            <h3>Next Up</h3>
            <div class="next-up-name">{nt_name}</div>
            <div class="next-up-date">{nt_date_display}</div>
            <div class="next-up-meta">{nt_type} &bull; {pts_line} &bull; Participation: {nt_part} pts</div>
        </div>"""

    return f"""
    <div class="overview-split">
        <div class="overview-left">
            <h2>Season at a Glance</h2>
            <div class="stat-grid">
                {stats_html}
            </div>
            <h2 class="section-h2">Participation &amp; Conditions</h2>
            {pc_html}
            {next_up_html}
        </div>
        <div class="overview-right">
            <h2>Top 15 Leaderboard</h2>
            <table class="top10-table">
                <thead>
                    <tr>
                        <th>Rank</th>
                        <th>+/-</th>
                        <th>Player</th>
                        <th>Events</th>
                        <th>Total</th>
                    </tr>
                </thead>
                <tbody>
                    {top15_rows}
                </tbody>
            </table>
        </div>
    </div>"""


def build_event_spotlight_html(player_data, player_events):
    """Build event spotlight for the most recent tournament.
    Shows teams as they appeared in Results - no splitting, no flight hierarchy."""
    played = [(i, t) for i, t in enumerate(TOURNAMENTS) if t[1] is not None]
    if not played:
        return ""
    idx, tourney = played[-1]
    display_name, sheet_name, event_type, team_size, places_paid, has_flights, date_str, _part_pts, _season = tourney

    if date_str:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        date_display = dt.strftime("%B %d, %Y")
    else:
        date_display = ""

    # Pull raw results from the sheet - keep teams together
    wb = openpyxl.load_workbook(RESULTS_FILE, data_only=True)
    ws = wb[sheet_name]

    # Collect all team results across all flights, flattened
    all_results = []  # (pos_val, team_str, score, to_par, per_player_pts, flight_num)

    if has_flights:
        flights = parse_flighted_sheet(ws)
        for flight_idx, flight_rows in enumerate(flights, start=1):
            flight_results = []
            for vals in flight_rows:
                pos_val = vals[1]
                players_str = vals[2]
                if pos_val is None or players_str is None:
                    continue
                players = split_players(str(players_str))
                flight_results.append((pos_val, players))

            pts = calc_points_for_flight(flight_results, event_type, places_paid)

            for vals in flight_rows:
                pos_val = vals[1]
                players_str = vals[2]
                to_par = vals[3] if len(vals) > 3 else ""
                score = vals[5] if len(vals) > 5 else (vals[4] if len(vals) > 4 else "")
                if pos_val is None or players_str is None:
                    continue
                # Get per-player pts (same for all players on the team)
                first_player = split_players(str(players_str))[0]
                earned = pts.get(first_player, 0)
                if earned > 0:
                    all_results.append((pos_val, str(players_str).strip(), score, to_par, earned, flight_idx))
    else:
        results = parse_single_flight_sheet(ws)
        pts = calc_points_for_flight(results, event_type, places_paid)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = list(row)
            if vals[0] and str(vals[0]).strip() in ("Pos.",):
                continue
            if vals[1] and str(vals[1]).strip() in ("Net",):
                continue
            if vals[0] and "golfgenius" in str(vals[0]):
                pos_val = vals[1]
                players_str = vals[2]
                to_par = vals[3] if len(vals) > 3 else ""
                score = vals[4] if len(vals) > 4 else ""
            elif vals[0] is not None:
                pos_val = vals[0]
                players_str = vals[1]
                to_par = vals[2] if len(vals) > 2 else ""
                score = vals[4] if len(vals) > 4 else ""
            else:
                continue
            if pos_val is None or players_str is None:
                continue
            try:
                parse_position(pos_val)
            except (ValueError, TypeError):
                continue
            first_player = split_players(str(players_str))[0]
            earned = pts.get(first_player, 0)
            if earned > 0:
                all_results.append((pos_val, str(players_str).strip(), score, to_par, earned, None))

    # Sort by points desc, then flight number (treats all flights equally)
    all_results.sort(key=lambda x: (-x[4], x[5] or 0, str(x[0])))

    # Count total participants
    total_teams = 0
    total_players_count = 0
    if has_flights:
        for flight_rows in flights:
            for vals in flight_rows:
                if vals[1] is not None and vals[2] is not None:
                    total_teams += 1
                    total_players_count += len(split_players(str(vals[2])))
    else:
        total_teams = len(results)
        total_players_count = sum(len(plist) for _, plist in results)

    # Build rows
    top_rows = ""
    for pos_val, team_str, score, to_par, earned, flight_num in all_results:
        pos_display = str(pos_val).strip()
        score_display = str(score) if score else ""
        pts_display = int(earned) if earned == int(earned) else f"{earned:.1f}"
        top_rows += f"""
                <tr>
                    <td class="rank">{pos_display}</td>
                    <td class="player">{team_str}</td>
                    <td class="pts">{score_display}</td>
                    <td class="total">{pts_display}</td>
                </tr>"""

    type_label = f"{team_size}-Man Event"
    teams_paid = len(all_results)
    flight_label = f"{len(flights)} flights" if has_flights else "No flights"
    ordinals = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th", 6: "6th", 7: "7th"}

    # Points structure line
    table_key = (event_type, places_paid)
    pts_table = POINTS_TABLE.get(table_key, {})
    structure_parts = [f"{ordinals.get(p, str(p))}: {int(v) if v == int(v) else v}" for p, v in sorted(pts_table.items())]
    structure_line = " | ".join(structure_parts)

    # Date with day abbreviation
    if date_str:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        end_str = MULTI_DAY_END.get(display_name)
        if end_str:
            dt_end = datetime.strptime(end_str, "%Y-%m-%d")
            if dt.month == dt_end.month:
                date_with_day = f"{dt.strftime('%B')} {dt.day}-{dt_end.day}, {dt.year}"
            else:
                date_with_day = f"{dt.strftime('%B')} {dt.day} - {dt_end.strftime('%B')} {dt_end.day}, {dt.year}"
        else:
            date_with_day = f"{dt.strftime('%B')} {dt.day}, {dt.year}"
    else:
        date_with_day = ""

    # Build small tables per flight showing paid places
    if has_flights:
        # Group results by flight
        from collections import defaultdict
        flight_results = defaultdict(list)
        for pos_val, team_str, score, to_par, earned, flight_num in all_results:
            flight_results[flight_num].append({
                "pos": str(pos_val).strip(),
                "player": team_str,
                "pts": earned
            })

        # Sort each flight by position
        for fnum in flight_results:
            flight_results[fnum].sort(key=lambda x: parse_position(x["pos"])[0])

        # Build tables grid
        tables_html = ""
        num_flights = len(flights)
        for fnum in sorted(flight_results.keys()):
            rows_html = ""
            for entry in flight_results[fnum]:
                pts_display = int(entry["pts"]) if entry["pts"] == int(entry["pts"]) else f"{entry['pts']:.1f}"
                rows_html += f"""
                        <tr>
                            <td class="flight-pos">{entry['pos']}</td>
                            <td class="flight-player">{entry['player']}</td>
                            <td class="flight-pts">{pts_display}</td>
                        </tr>"""
            tables_html += f"""
                <div class="flight-table-wrap">
                    <table class="flight-table">
                        <thead>
                            <tr><th colspan="3">Flight {fnum}</th></tr>
                            <tr class="flight-subhead"><th>Place</th><th>Player</th><th>Pts</th></tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>
                </div>"""
        results_section = f'<div class="flight-tables-grid">{tables_html}</div>'
    else:
        # Non-flighted: single table
        rows_html = ""
        sorted_results = sorted(all_results, key=lambda x: parse_position(x[0])[0])
        for pos_val, team_str, score, to_par, earned, flight_num in sorted_results:
            pts_display = int(earned) if earned == int(earned) else f"{earned:.1f}"
            rows_html += f"""
                    <tr>
                        <td class="flight-pos">{str(pos_val).strip()}</td>
                        <td class="flight-player">{team_str}</td>
                        <td class="flight-pts">{pts_display}</td>
                    </tr>"""
        results_section = f"""
            <div class="flight-tables-grid single">
                <div class="flight-table-wrap">
                    <table class="flight-table">
                        <thead>
                            <tr class="flight-subhead"><th>Place</th><th>Player</th><th>Pts</th></tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>
                </div>
            </div>"""

    return f"""
    <div class="spotlight-info">
        <div class="spotlight-date">{date_with_day}</div>
        <div class="spotlight-meta">{type_label} &bull; {flight_label} &bull; {total_teams} Players &bull; {places_paid} Places Paid per Flight</div>
        <div class="spotlight-structure">Per-player points: {structure_line}</div>
        <p class="spotlight-note">{teams_paid} players earned placement points &bull; All flights weighted equally</p>
    </div>
    {results_section}"""


def build_results_html():
    """Build full tournament results pages from the Results.xlsx data."""
    wb = openpyxl.load_workbook(RESULTS_FILE, data_only=True)
    sections = ""

    for display_name, sheet_name, event_type, team_size, places_paid, has_flights, date_str, _part_pts, _season in TOURNAMENTS:
        if sheet_name is None or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_display = dt.strftime("%B %d, %Y")
        else:
            date_display = ""

        type_label = f"{team_size}-Man Event"
        flight_label = "Flighted" if has_flights else "Non-Flighted"

        # Detect Member-Member match play format
        g1 = ws.cell(1, 7).value
        if g1 and str(g1).strip() == "Results":
            all_participants, pts = parse_member_member_sheet(ws, event_type, places_paid)
            # Build results table from G & H columns
            rows_html = ""
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=8, values_only=True):
                place_val, player_name = row
                if place_val is None or player_name is None or not isinstance(place_val, (int, float)):
                    continue
                ordinal = {1: "1st", 2: "2nd", 3: "3rd"}.get(int(place_val), str(int(place_val)))
                # Find canonical name and their earned points
                pname = str(player_name).strip()
                name_parts = pname.rsplit(None, 1)
                if len(name_parts) == 2:
                    converted = f"{name_parts[1]}, {name_parts[0]}"
                else:
                    converted = pname
                earned = pts.get(converted, 0)
                # Try last-name match
                if earned == 0:
                    for canon_name, canon_pts in pts.items():
                        if canon_name.split(',')[0].strip().lower() == (name_parts[1].lower() if len(name_parts) == 2 else ''):
                            earned = canon_pts
                            break
                pts_display = int(earned) if earned == int(earned) else f"{earned:.1f}"
                rows_html += f"""
                        <tr>
                            <td class="rank">{ordinal}</td>
                            <td class="player">{pname}</td>
                            <td class="pts">{pts_display}</td>
                            <td class="pts"></td>
                        </tr>"""
            content = f"""
                    <table class="results-table">
                        <thead>
                            <tr>
                                <th>Place</th>
                                <th>Player</th>
                                <th>Points</th>
                                <th></th>
                            </tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>"""

            sections += f"""
        <div class="results-tourney">
            <div class="results-header">
                <h3>{display_name}</h3>
                <span class="results-meta">{date_display} &bull; {type_label} &bull; 5 Flights (Match Play) &bull; {places_paid} Places Paid</span>
            </div>
            {content}
        </div>"""
            continue

        if has_flights:
            flights = parse_flighted_sheet(ws)
            # Detect Member-Member-style layout: a "Team" column plus a cumulative
            # "Points" total column (no To Par / Net). Render Pos | Team | Points.
            mm_points = False
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if (len(row) >= 9 and row[8] and str(row[8]).strip().lower() == "points"
                        and row[2] and str(row[2]).strip().lower() == "team"):
                    mm_points = True
                    break
            flights_html = ""
            for fi, flight_rows in enumerate(flights, 1):
                rows_html = ""
                for vals in flight_rows:
                    pos_val = vals[1]
                    players_str = vals[2]
                    if pos_val is None or players_str is None:
                        continue
                    pos_display = str(pos_val).strip()
                    if mm_points:
                        total = vals[8] if len(vals) > 8 and vals[8] is not None else ""
                        rows_html += f"""
                        <tr>
                            <td class="rank">{pos_display}</td>
                            <td class="player">{players_str}</td>
                            <td class="pts">{total}</td>
                        </tr>"""
                    else:
                        score = vals[5] if len(vals) > 5 else vals[4] if len(vals) > 4 else ""
                        to_par = vals[3] if len(vals) > 3 else ""
                        to_par_display = str(to_par) if to_par is not None else ""
                        score_display = str(score) if score is not None else ""
                        rows_html += f"""
                        <tr>
                            <td class="rank">{pos_display}</td>
                            <td class="player">{players_str}</td>
                            <td class="pts">{to_par_display}</td>
                            <td class="pts">{score_display}</td>
                        </tr>"""
                if mm_points:
                    head = "<th>Pos</th><th>Players</th><th>Points</th>"
                else:
                    head = "<th>Pos</th><th>Players</th><th>To Par</th><th>Net</th>"
                flights_html += f"""
                    <h4 class="flight-header">Flight {fi}</h4>
                    <table class="results-table">
                        <thead>
                            <tr>{head}</tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>"""
            content = flights_html
        else:
            results = parse_single_flight_sheet(ws)
            rows_html = ""
            # Re-parse to get scores
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                vals = list(row)
                if vals[0] and str(vals[0]).strip() in ("Pos.",):
                    continue
                if vals[1] and str(vals[1]).strip() in ("Net",):
                    continue
                if vals[0] and "golfgenius" in str(vals[0]):
                    pos_val = vals[1]
                    players_str = vals[2]
                    to_par = vals[3] if len(vals) > 3 else ""
                    score = vals[4] if len(vals) > 4 else ""
                elif vals[0] is not None:
                    pos_val = vals[0]
                    players_str = vals[1]
                    to_par = vals[2] if len(vals) > 2 else ""
                    score = vals[4] if len(vals) > 4 else ""
                else:
                    continue
                if pos_val is None or players_str is None:
                    continue
                try:
                    parse_position(pos_val)
                except (ValueError, TypeError):
                    continue
                pos_display = str(pos_val).strip()
                to_par_display = str(to_par) if to_par is not None else ""
                score_display = str(score) if score is not None else ""
                rows_html += f"""
                        <tr>
                            <td class="rank">{pos_display}</td>
                            <td class="player">{players_str}</td>
                            <td class="pts">{to_par_display}</td>
                            <td class="pts">{score_display}</td>
                        </tr>"""
            content = f"""
                    <table class="results-table">
                        <thead>
                            <tr>
                                <th>Pos</th>
                                <th>Players</th>
                                <th>To Par</th>
                                <th>Net</th>
                            </tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>"""

        sections += f"""
        <div class="results-tourney">
            <div class="results-header">
                <h3>{display_name}</h3>
                <span class="results-meta">{date_display} &bull; {type_label} &bull; {flight_label} &bull; {places_paid} Places Paid</span>
            </div>
            {content}
        </div>"""

    return sections


def get_field_stats():
    """Analyze each tournament's field size from the results workbook."""
    wb = openpyxl.load_workbook(RESULTS_FILE, data_only=True)
    stats = {}

    for display_name, sheet_name, event_type, team_size, places_paid, has_flights, date_str, _part_pts, _season in TOURNAMENTS:
        if sheet_name is None or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Detect Member-Member match play format
        g1 = ws.cell(1, 7).value
        if g1 and str(g1).strip() == "Results":
            all_participants, _ = parse_member_member_sheet(ws, event_type, places_paid)
            total_players = len(all_participants)
            total_teams = total_players // team_size if team_size > 0 else total_players
            stats[display_name] = {
                "total_teams": total_teams,
                "total_players": total_players,
                "num_flights": 5,
                "flight_sizes": [6] * 5,
                "places_per_flight": places_paid,
                "total_places_paid": places_paid * 5,
                "pct_field": (places_paid * 5 / total_teams * 100) if total_teams else 0,
            }
            continue

        if has_flights:
            flights = parse_flighted_sheet(ws)
            flight_sizes = []
            total_teams = 0
            total_players = 0
            for flight_rows in flights:
                teams_in_flight = 0
                for vals in flight_rows:
                    if vals[1] is not None and vals[2] is not None:
                        teams_in_flight += 1
                        players_in_team = len(split_players(str(vals[2])))
                        total_players += players_in_team
                flight_sizes.append(teams_in_flight)
                total_teams += teams_in_flight
            num_flights = len(flights)
            # Places paid is per flight
            total_places_paid = places_paid * num_flights
            pct = (total_places_paid / total_teams * 100) if total_teams else 0
            stats[display_name] = {
                "total_teams": total_teams,
                "total_players": total_players,
                "num_flights": num_flights,
                "flight_sizes": flight_sizes,
                "places_per_flight": places_paid,
                "total_places_paid": total_places_paid,
                "pct_field": pct,
            }
        else:
            results = parse_single_flight_sheet(ws)
            total_teams = len(results)
            total_players = sum(len(plist) for _, plist in results)
            pct = (places_paid / total_teams * 100) if total_teams else 0
            stats[display_name] = {
                "total_teams": total_teams,
                "total_players": total_players,
                "num_flights": 0,
                "flight_sizes": [],
                "places_per_flight": places_paid,
                "total_places_paid": places_paid,
                "pct_field": pct,
            }

    return stats


def build_appendix_html():
    """Build HTML for the scoring appendix in a two-column grid layout with field analysis."""
    ordinals = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th", 6: "6th", 7: "7th"}
    field_stats = get_field_stats()
    cards = []

    for display_name, sheet_name, event_type, team_size, places_paid, has_flights, date_str, _part_pts, _season in TOURNAMENTS:
        if sheet_name is None:
            continue

        table_key = (event_type, places_paid)
        per_player_table = POINTS_TABLE.get(table_key, {})
        if not per_player_table:
            continue

        type_label = f"{team_size}-Man Event"
        flight_label = "Flighted" if has_flights else "Non-Flighted"

        # Format date
        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_display = dt.strftime("%B %d, %Y")
        else:
            date_display = "TBD"

        # Field stats
        fs = field_stats.get(display_name, {})
        total_teams = fs.get("total_teams", 0)
        total_players = fs.get("total_players", 0)
        num_flights = fs.get("num_flights", 0)
        total_places_paid = fs.get("total_places_paid", 0)
        pct = fs.get("pct_field", 0)

        if has_flights:
            flight_sizes = fs.get("flight_sizes", [])
            field_line = f"{total_teams} teams ({total_players} players) across {num_flights} flights"
            payout_line = f"{places_paid} places paid per flight - {total_places_paid} of {total_teams} teams ({pct:.0f}% of field)"
        else:
            field_line = f"{total_teams} teams ({total_players} players)"
            payout_line = f"{total_places_paid} of {total_teams} teams paid ({pct:.0f}% of field)"

        # Build points rows
        rows = ""
        for place in sorted(per_player_table.keys()):
            per_player = per_player_table[place]
            team_pts = per_player * team_size
            team_display = int(team_pts) if team_pts == int(team_pts) else f"{team_pts:.1f}"
            player_display = int(per_player) if per_player == int(per_player) else f"{per_player:.1f}"
            rows += f"""
                    <tr>
                        <td>{ordinals.get(place, str(place))}</td>
                        <td>{team_display}</td>
                        <td>{player_display}</td>
                    </tr>"""

        card = f"""
            <div class="appendix-card">
                <h3>{display_name}</h3>
                <div class="tourney-date">{date_display}</div>
                <div class="tourney-meta">{type_label} &bull; {flight_label}</div>
                <div class="tourney-field">{field_line}</div>
                <div class="tourney-payout">{payout_line}</div>
                <table>
                    <thead>
                        <tr>
                            <th>Place</th>
                            <th>Team Pts</th>
                            <th>Per Player</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows}
                    </tbody>
                </table>
            </div>"""
        cards.append(card)

    grid = f"""
        <div class="appendix-grid">
            {''.join(cards)}
        </div>
        <div class="appendix-footer">
            Participation points vary by event (see Schedule).
        </div>"""
    return grid


def build_season_schedule_html():
    """Build the season schedule page with Completed and Upcoming segments."""
    today = datetime.now().strftime("%Y-%m-%d")

    # Find next up tournament
    next_up_name = None
    for t in TOURNAMENTS:
        if t[1] is None and t[6] and t[6] >= today and t[0] not in CANCELLED_EVENTS:
            next_up_name = t[0]
            break

    # Gather field stats for completed tournaments
    field_stats = get_field_stats()

    # Helper for date display
    def fmt_date(name, date_str):
        if not date_str:
            return "TBD"
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        end_str = MULTI_DAY_END.get(name)
        if end_str:
            dt_end = datetime.strptime(end_str, "%Y-%m-%d")
            if dt.month == dt_end.month:
                return f"{dt.strftime('%b')} {dt.day}-{dt_end.day}"
            return f"{dt.strftime('%b %d')} - {dt_end.strftime('%b %d')}"
        return dt.strftime("%b %d")

    # Helper for format label
    def fmt_type(team_size):
        if team_size == 1:
            return "Individual"
        elif team_size > 1:
            return f"{team_size}-Man"
        return "-"

    # Helper for 1st place points (always per-player, no caveats)
    def fmt_first(name, etype, places, team_size):
        if "Lonely Guy" in name:
            return "350"
        if "2 Man Match Play" in name:
            return "175"
        pts_table = POINTS_TABLE.get((etype, places), {})
        first_pp = pts_table.get(1, 0)
        if first_pp > 0:
            return str(int(first_pp))
        elif name == "Eldo Cup":
            return "-"
        return "-"

    # ── Split into completed, on-going, cancelled, and upcoming ──
    completed = []
    ongoing = []
    cancelled = []
    upcoming = []
    for t in TOURNAMENTS:
        name, sheet, etype, team_size, places, has_flights, date_str, part_pts, season = t
        if name in CANCELLED_EVENTS:
            cancelled.append(t)
        elif etype in MATCH_PLAY_TYPES:
            ongoing.append(t)
        elif sheet is not None:
            completed.append(t)
        else:
            upcoming.append(t)

    # Single table with section header rows - guarantees perfect column alignment

    def sched_row(date_display, name, fmt_label, first_str, part_str, status_cls, status_text, hide_date=False):
        date_style = ' style="color: transparent;"' if hide_date else ''
        return f"""
                <tr>
                    <td class="sched-date"{date_style}>{date_display}</td>
                    <td class="sched-name">{name}</td>
                    <td class="sched-format">{fmt_label}</td>
                    <td class="sched-pts">{first_str}</td>
                    <td class="sched-pts">{part_str}</td>
                    <td class="sched-status"><span class="{status_cls}">{status_text}</span></td>
                </tr>"""

    def section_header(title):
        return f"""
                <tr class="sched-section-row">
                    <td colspan="6" class="sched-section">{title}</td>
                </tr>"""

    all_rows = section_header("Completed")

    # ── Completed rows (merge completed + cancelled, sort by date) ──
    completed_all = completed + cancelled
    completed_all.sort(key=lambda t: t[6] or "9999-99-99")  # sort by date_str
    for t in completed_all:
        name, sheet, etype, team_size, places, has_flights, date_str, part_pts, season = t
        if name in CANCELLED_EVENTS:
            all_rows += sched_row(
                fmt_date(name, date_str), name, fmt_type(team_size),
                fmt_first(name, etype, places, team_size), "",
                "status-cancelled", "Cancelled - All Players Qualify")
        else:
            all_rows += sched_row(
                fmt_date(name, date_str), name, fmt_type(team_size),
                fmt_first(name, etype, places, team_size), str(part_pts),
                "status-played", "Completed")

    # ── Match-play rounds 1 & 2 (complete and scored) ──
    for t in ongoing:
        info = MATCH_PLAY_ROUNDS.get(t[0])
        if not info:
            continue
        all_rows += sched_row(
            info["done_dates"], f'{info["label"]} - {info["done_label"]}',
            fmt_type(info["team_size"]), info["done_best"], info["part"],
            "status-played", "Completed")

    # ── Ongoing rows (next match-play round to be played) ──
    if ongoing:
        all_rows += section_header("Ongoing")
        for t in ongoing:
            info = MATCH_PLAY_ROUNDS.get(t[0])
            if not info:
                continue
            all_rows += sched_row(
                "", f'{info["label"]} - {info["next_round"]}',
                fmt_type(info["team_size"]), "-", "-",
                "status-next", "Next Round", hide_date=True)
        all_rows += f"""
                <tr class="sched-note-row">
                    <td colspan="6" class="sched-note">Rounds 1 through 4 are complete and scored (participation plus points per round won). Points for the remaining rounds will be added as they are played.</td>
                </tr>"""

    # ── Upcoming rows ──
    all_rows += section_header("Upcoming")
    for t in upcoming:
        name, sheet, etype, team_size, places, has_flights, date_str, part_pts, season = t
        if name == next_up_name:
            s_cls, s_txt = "status-next", "Next Up"
        else:
            note = SCHEDULE_NOTES.get(name)
            if note:
                s_cls, s_txt = "status-note", note
            else:
                s_cls, s_txt = "status-upcoming", "Upcoming"
        all_rows += sched_row(
            fmt_date(name, date_str), name, fmt_type(team_size),
            fmt_first(name, etype, places, team_size), str(part_pts),
            s_cls, s_txt)

    html = f"""
            <table class="schedule-table">
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Tournament</th>
                        <th>Format</th>
                        <th>1st Place</th>
                        <th>Part Pts</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>{all_rows}</tbody>
            </table>"""

    return html


def build_standings_snapshot(player_data, player_events, through_idx):
    """Build standings as if only tournaments 0..through_idx existed."""
    included = [TOURNAMENTS[i][0] for i in range(through_idx + 1) if TOURNAMENTS[i][1] is not None]
    part_pts_map = {t[0]: t[7] for t in TOURNAMENTS}
    scores = {}
    for player, tourney_pts in player_data.items():
        total = sum(tourney_pts.get(t, 0) for t in included)
        participated = [t for t in included if t in player_events.get(player, set())]
        events = len(participated)
        total += sum(part_pts_map.get(t, PARTICIPATION_PTS) for t in participated)
        if total > 0 or events > 0:
            scores[player] = total
    sorted_players = sorted(scores.items(), key=lambda x: (-x[1], x[0]))
    ranks = {}
    for i, (player, pts) in enumerate(sorted_players):
        if i == 0:
            ranks[player] = (1, pts)
        elif pts == sorted_players[i - 1][1]:
            ranks[player] = (ranks[sorted_players[i - 1][0]][0], pts)
        else:
            ranks[player] = (i + 1, pts)
    return ranks


def calc_rank_changes(standings, player_data, player_events):
    """Calculate rank change for each player vs. prior tournament snapshot."""
    played_indices = [i for i, t in enumerate(TOURNAMENTS) if t[1] is not None]
    if len(played_indices) < 2:
        return {entry["player"]: None for entry in standings}

    prev_idx = played_indices[-2]
    before = build_standings_snapshot(player_data, player_events, prev_idx)

    field_size = len(standings)
    changes = {}
    for entry in standings:
        player = entry["player"]
        new_rank = entry["rank"]
        if player in before:
            old_rank = before[player][0]
            changes[player] = old_rank - new_rank
        else:
            changes[player] = None  # new entrant
    return changes


def generate_html(standings, tournament_names, player_data, player_events):
    """Generate standings HTML file."""
    now = datetime.now().strftime("%B %d, %Y %I:%M %p")
    rank_changes = calc_rank_changes(standings, player_data, player_events)

    # Find latest event name for spotlight
    played = [(i, t) for i, t in enumerate(TOURNAMENTS) if t[1] is not None]
    latest_event_name = played[-1][1][0] if played else ""

    tourney_headers = "".join(f"<th>{t}</th>" for t in tournament_names)

    # All players in one table, ranked
    placement_standings = standings

    part_pts_map = {t[0]: t[7] for t in TOURNAMENTS}
    event_order = [t[0] for t in TOURNAMENTS]

    rows_html = ""
    for entry in placement_standings:
        rank = entry["rank"]
        # Only show "T" prefix if the tie exists within the placement table, not just the full standings
        rank_display = f"T{rank}" if any(
            e["rank"] == rank and e["player"] != entry["player"] for e in placement_standings
        ) else str(rank)

        tourney_cells = ""
        for t in tournament_names:
            pts = entry["tournaments"][t]
            if pts > 0:
                display = int(pts) if pts == int(pts) else f"{pts:.1f}"
                tourney_cells += f'<td class="pts">{display}</td>'
            else:
                tourney_cells += '<td class="pts empty">-</td>'

        total_display = int(entry["total"]) if entry["total"] == int(entry["total"]) else f"{entry['total']:.1f}"
        part_display = int(entry["participation"]) if entry["participation"] == int(entry["participation"]) else f"{entry['participation']:.1f}"

        delta_cell = format_rank_delta(rank_changes.get(entry["player"]))

        # Participation tooltip: checklist of events the player entered
        played = player_events.get(entry["player"], set())
        tip_items = "".join(
            f'<span class="pt-item"><span class="pt-chk">&#10003;</span>{ev}'
            f'<span class="pt-pp">+{int(part_pts_map.get(ev, PARTICIPATION_PTS))}</span></span>'
            for ev in event_order if ev in played
        )
        if tip_items:
            part_cell = (f'<td class="pts part-cell">{part_display}'
                         f'<span class="part-tip"><span class="part-tip-h">Events Entered</span>'
                         f'{tip_items}</span></td>')
        else:
            part_cell = f'<td class="pts">{part_display}</td>'

        rows_html += f"""
        <tr>
            <td class="rank">{rank_display}</td>
            {delta_cell}
            <td class="player">{entry['player']}</td>
            {tourney_cells}
            <td class="events">{entry['events']}</td>
            {part_cell}
            <td class="total">{total_display}</td>
        </tr>"""

    # Build grouped section for participation-only players
    # Build participation points reference for footer
    played_tournaments = [(t[0], t[7]) for t in TOURNAMENTS if event_has_data(t)]
    part_by_value = {}
    for tname, ppts in played_tournaments:
        part_by_value.setdefault(ppts, []).append(tname)
    part_ref_parts = []
    for ppts in sorted(part_by_value.keys(), reverse=True):
        tnames_list = ", ".join(part_by_value[ppts])
        part_ref_parts.append(f"({int(ppts)}) {tnames_list}")
    part_reference = " &bull; ".join(part_ref_parts)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>MGA Ryder Cup Points Standings 2025-26</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        background: #f0f2f5;
        color: #1a1a2e;
        padding: 20px;
    }}
    .container {{
        max-width: 1200px;
        margin: 0 auto;
        background: #fff;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        overflow: hidden;
    }}
    .header {{
        background: #fff;
        color: #1a472a;
        padding: 24px 40px;
        text-align: center;
        border-bottom: 3px solid #1a472a;
    }}
    .header h1 {{
        font-size: 28px;
        font-weight: 800;
        letter-spacing: 1px;
        margin-bottom: 4px;
        color: #1a472a;
    }}
    .header .subtitle {{
        font-size: 14px;
        color: #333;
        font-weight: 500;
    }}
    .table-wrapper {{
        overflow-x: auto;
        padding: 0;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 14px;
    }}
    thead th {{
        background: #f5f5f5;
        color: #1a472a;
        padding: 12px 10px;
        text-align: center;
        font-weight: 700;
        font-size: 12px;
        border-bottom: 2px solid #1a472a;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        white-space: normal;
        min-width: 50px;
        position: sticky;
        top: 0;
    }}
    tbody tr {{
        border-bottom: 1px solid #e8e8e8;
        transition: background 0.15s;
    }}
    tbody tr:hover {{ background: #f5faf7; }}
    tbody tr:nth-child(even) {{ background: #fafbfc; }}
    tbody tr:nth-child(even):hover {{ background: #f0f7f2; }}
    td {{
        padding: 10px 10px;
        text-align: center;
    }}
    td.rank {{ font-weight: 700; color: #264636; }}
    td.rank-delta {{
        font-weight: 600; font-size: 12px;
        font-variant-numeric: tabular-nums;
        text-align: center; white-space: nowrap;
    }}
    td.rank-delta.up   {{ color: #1a7f3e; }}
    td.rank-delta.down {{ color: #b3261e; }}
    td.rank-delta.flat {{ color: #aaa; }}
    td.rank-delta.new  {{ color: #1a472a; font-size: 10px; letter-spacing: 0.5px; }}
    td.player {{ text-align: left; font-weight: 500; white-space: nowrap; }}
    td.pts {{ font-variant-numeric: tabular-nums; }}
    td.empty {{ color: #ccc; }}
    td.events {{ font-weight: 600; color: #555; }}
    td.total {{
        font-weight: 700; color: #1a472a;
        font-size: 15px; background: #f0f7f2;
    }}
    .footer {{
        text-align: center; padding: 15px;
        font-size: 12px; color: #888;
        border-top: 1px solid #e8e8e8;
    }}

    /* ── Overview ── */
    .overview-split {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 30px;
        padding: 30px 40px 30px;
    }}
    .overview-left h2, .overview-right h2 {{
        font-size: 18px; color: #1a472a;
        margin-bottom: 14px;
        border-bottom: 2px solid #dde3da;
        padding-bottom: 6px;
    }}
    .stat-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr 1fr;
        gap: 10px;
    }}
    .stat-card {{
        background: #f0f7f2; border-radius: 8px;
        padding: 12px 8px; text-align: center;
    }}
    .stat-number {{
        font-size: 26px; font-weight: 700; color: #1a472a;
    }}
    .stat-label {{
        font-size: 10px; text-transform: uppercase;
        color: #666; letter-spacing: 0.5px; margin-top: 2px;
    }}
    .top10-table {{ width: 100%; font-size: 13px; }}
    .top10-table thead th {{ padding: 8px 10px; font-size: 11px; }}
    .top10-table tbody td {{ padding: 6px 10px; }}
    .section-h2 {{
        font-size: 16px; color: #1a472a;
        margin: 16px 0 10px;
        border-bottom: 2px solid #dde3da;
        padding-bottom: 6px;
    }}
    /* ── Forecast Cards ── */
    .fc-container {{
        display: flex;
        gap: 12px;
        justify-content: center;
        margin-top: 8px;
    }}
    .fc-card {{
        flex: 1;
        max-width: 140px;
        text-align: center;
        padding: 14px 8px;
        background: #fafbfc;
        border: 1px solid #e0e4dd;
        border-radius: 8px;
    }}
    .fc-name {{
        font-size: 11px;
        font-weight: 700;
        color: #1a472a;
        margin-bottom: 2px;
        letter-spacing: 0.3px;
        min-height: 26px;
        line-height: 1.2;
    }}
    .fc-date {{
        font-size: 10px;
        color: #888;
        margin-bottom: 8px;
    }}
    .fc-field {{
        font-size: 24px;
        font-weight: 700;
        color: #1a472a;
        line-height: 1;
    }}
    .fc-field-label {{
        font-size: 8px;
        text-transform: uppercase;
        letter-spacing: 1px;
        color: #999;
        margin-bottom: 8px;
    }}
    .fc-icon {{
        margin: 4px 0;
    }}
    .fc-temps {{
        font-size: 11px;
        font-weight: 600;
        color: #1a1a2e;
    }}
    .fc-wind {{
        font-size: 9px;
        color: #888;
    }}
    .next-up {{
        margin-top: 16px;
        padding: 14px 16px;
        background: #f0f7f2;
        border-radius: 8px;
        border-left: 4px solid #1a472a;
    }}
    .next-up h3 {{
        font-size: 11px; text-transform: uppercase;
        color: #666; letter-spacing: 1px; margin-bottom: 4px;
    }}
    .next-up-name {{
        font-size: 18px; font-weight: 700; color: #1a472a;
        margin-bottom: 2px;
    }}
    .next-up-meta {{
        font-size: 12px; color: #555;
    }}

    /* ── Event Spotlight ── */
    .spotlight-container {{ page-break-before: always; }}
    .spotlight-info {{ padding: 16px 30px 0; }}
    .spotlight-info h3 {{
        font-size: 14px; color: #1a472a; margin-bottom: 4px;
    }}
    .spotlight-meta {{
        font-size: 11px; color: #666; margin-bottom: 6px;
    }}
    .spotlight-note {{
        font-size: 11px; color: #888; margin-bottom: 6px;
    }}
    .spotlight-date {{
        font-size: 14px;
        font-weight: 700;
        color: #1a472a;
        margin-bottom: 2px;
    }}
    .spotlight-structure {{
        font-size: 11px;
        color: #264636;
        font-weight: 600;
        margin: 4px 0;
        padding: 4px 10px;
        background: #f0f7f2;
        border-radius: 4px;
        display: inline-block;
    }}
    .flight-tables-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        gap: 16px;
        padding: 12px 30px 20px;
    }}
    .flight-tables-grid.single {{
        grid-template-columns: minmax(200px, 300px);
        justify-content: center;
    }}
    .flight-table-wrap {{
        background: #fafbfc;
        border: 1px solid #e8e8e8;
        border-radius: 6px;
        overflow: hidden;
    }}
    .flight-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 11px;
    }}
    .flight-table thead th {{
        background: #1a472a;
        color: white;
        font-weight: 600;
        padding: 6px 8px;
        text-align: center;
    }}
    .flight-table .flight-subhead th {{
        background: #f5f5f5;
        color: #333;
        font-weight: 500;
        font-size: 9px;
        padding: 4px 8px;
        border-bottom: 1px solid #ddd;
    }}
    .flight-table tbody tr:nth-child(even) {{
        background: #f9f9f9;
    }}
    .flight-table td {{
        padding: 5px 8px;
        border-bottom: 1px solid #eee;
    }}
    .flight-pos {{
        text-align: center;
        font-weight: 500;
        width: 40px;
    }}
    .flight-player {{
        text-align: left;
    }}
    .flight-pts {{
        text-align: center;
        font-weight: 600;
        color: #1a472a;
        width: 40px;
    }}

    /* ── Full Standings ── */
    /* overflow:visible (not auto) so the column headers stay sticky to the
       top of the page as the long table scrolls. On narrow screens the table
       collapses to four columns (see the screen media query) so it never
       overflows the viewport. */
    .standings-container {{ page-break-before: always; overflow: visible; }}
    .standings-container .table-wrapper {{ overflow: visible; }}
    .standings-container thead th {{
        position: sticky;
        top: 0;
        background: #f5f5f5;
        z-index: 10;
        box-shadow: 0 1px 0 #1a472a;
    }}
    .standings-footer {{
        text-align: right;
        font-size: 11px;
        color: #999;
        padding: 8px 4px;
        border-top: 1px solid #e8e8e8;
    }}

    /* ── Appendix ── */
    .appendix-container {{ page-break-before: always; }}
    .appendix {{ padding: 20px 30px; }}
    .appendix-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 20px;
    }}
    .appendix-card {{
        border: 1px solid #dde3da; border-radius: 8px;
        padding: 16px; background: #fafbfc;
    }}
    .appendix-card {{ page-break-inside: avoid; padding: 12px; }}
    .appendix-card h3 {{ font-size: 14px; color: #1a472a; margin-bottom: 2px; }}
    .appendix-card .tourney-date {{
        font-size: 11px; color: #1a472a; font-weight: 600; margin-bottom: 2px;
    }}
    .appendix-card .tourney-meta {{ font-size: 10px; color: #666; margin-bottom: 2px; }}
    .appendix-card .tourney-field {{ font-size: 10px; color: #444; margin-bottom: 2px; }}
    .appendix-card .tourney-payout {{
        font-size: 10px; color: #264636; font-weight: 600; margin-bottom: 6px;
    }}
    .appendix-card table {{ width: 100%; margin-bottom: 0; }}
    .appendix-card thead th {{ padding: 3px 6px; font-size: 9px; }}
    .appendix-card tbody td {{ padding: 2px 6px; font-size: 11px; }}
    .appendix-footer {{
        margin-top: 12px; padding: 8px 16px;
        background: #f0f7f2; border-radius: 6px;
        font-size: 11px; color: #264636;
        font-weight: 500; text-align: center;
    }}

    /* ── Season Schedule ── */
    .schedule-container {{ page-break-before: always; page-break-inside: avoid; }}
    .schedule {{ padding: 12px 30px 14px; }}
    .season-block {{ margin-bottom: 8px; }}
    .season-heading {{
        font-size: 14px; color: #1a472a;
        margin-bottom: 4px;
        border-bottom: 2px solid #dde3da;
        padding-bottom: 2px;
    }}
    .schedule-table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
    .schedule-table thead th {{
        background: #f5f5f5; color: #1a472a;
        padding: 6px 8px; text-align: center;
        font-size: 9px; text-transform: uppercase;
        font-weight: 700; letter-spacing: 0.5px;
        border-bottom: 2px solid #1a472a;
    }}
    .schedule-table tbody tr {{
        border-bottom: 1px solid #e8e8e8;
    }}
    .schedule-table tbody tr:nth-child(even) {{ background: #fafbfc; }}
    .schedule-table td {{ padding: 5px 8px; text-align: center; vertical-align: middle; }}
    .sched-date {{ font-weight: 600; color: #264636; white-space: nowrap; }}
    .sched-name {{ font-weight: 600; color: #1a472a; }}
    .sched-format {{ color: #555; }}
    .sched-pts {{ font-size: 10px; color: #444; }}
    .sched-part {{ font-size: 10px; color: #444; }}
    .sched-section-row td {{
        background: #fff;
        border-bottom: 2px solid #dde3da;
        padding-top: 16px;
    }}
    .sched-section {{
        font-size: 14px;
        font-weight: 700;
        color: #1a472a;
        text-align: left;
    }}
    .sched-section-row:first-child td {{
        padding-top: 4px;
    }}
    .sched-note-row td {{
        background: #fff;
        border-bottom: none;
    }}
    .sched-note {{
        font-size: 10px;
        color: #264636;
        font-weight: 500;
        text-align: left;
        padding: 6px 8px 12px;
    }}
    .status-played {{
        background: #e8f5e9; color: #2e7d32; padding: 1px 6px;
        border-radius: 10px; font-size: 9px; font-weight: 600;
    }}
    .status-next {{
        background: #fff3e0; color: #e65100; padding: 1px 6px;
        border-radius: 10px; font-size: 9px; font-weight: 600;
    }}
    .status-upcoming {{
        background: #e3f2fd; color: #1565c0; padding: 1px 6px;
        border-radius: 10px; font-size: 9px; font-weight: 600;
    }}
    .status-note {{
        background: #f5f5f5; color: #666; padding: 1px 6px;
        border-radius: 10px; font-size: 9px; font-weight: 600;
    }}
    .status-cancelled {{
        background: #fff8e1; color: #f9a825; padding: 1px 6px;
        border-radius: 10px; font-size: 9px; font-weight: 600;
    }}
    .next-up-date {{
        font-size: 14px; font-weight: 600; color: #264636;
        margin-bottom: 4px;
    }}
    .schedule-footer {{
        margin-top: 10px; padding: 8px 16px;
        background: #f0f7f2; border-radius: 6px;
        font-size: 10px; color: #264636;
        font-weight: 500; text-align: center;
    }}
    /* ── Grouped Standings ── */
    .compact-group-row td {{
        border-top: 2px solid #dde3da;
        vertical-align: top;
        padding-top: 12px;
        padding-bottom: 12px;
    }}
    .compact-names {{
        white-space: normal !important;
        line-height: 1.8;
        font-size: 13px;
    }}

    /* ── Rotate hint (mobile portrait only) ── */
    .rotate-hint {{
        display: none;
        align-items: center;
        justify-content: center;
        gap: 10px;
        padding: 12px 16px;
        margin: 0 0 12px;
        background: #fff3e0;
        border: 1px solid #ffe0b2;
        border-radius: 8px;
        color: #e65100;
        font-size: 13px;
        font-weight: 600;
    }}
    .rotate-icon {{
        width: 28px;
        height: 28px;
        flex-shrink: 0;
    }}

    /* ── Mobile standings: tap-to-expand (defaults; activated under media query) ── */
    .detail-row {{ display: none; }}
    .detail-caret {{ display: none; }}
    .mobile-tip {{ display: none; }}

    @media print {{
        body {{ background: #fff; padding: 0; }}
        .container {{ box-shadow: none; border-radius: 0; }}
        .header {{ padding: 20px; }}
        tbody tr:nth-child(even) {{ background: #f5f5f5; }}
        .rotate-hint {{ display: none !important; }}
        /* Print keeps the full wide table with sticky headers */
        .standings-container,
        .standings-container .table-wrapper {{ overflow: visible; }}
    }}
    @media (max-width: 768px) {{
        body {{ padding: 8px; }}
        .container {{ border-radius: 8px; }}
        .overview-split {{ grid-template-columns: 1fr; padding: 16px; gap: 16px; }}
        .appendix-grid {{ grid-template-columns: 1fr; }}
        .header {{ padding: 16px; }}
        .header h1 {{ font-size: 18px; }}
        .header .subtitle {{ font-size: 12px; }}
        table {{ font-size: 11px; }}
        td, th {{ padding: 6px 4px; }}
        .schedule {{ padding: 10px 16px; }}
        .appendix {{ padding: 16px; }}
        .spotlight-info {{ padding: 12px 16px 0; }}
        .flight-tables-grid {{ padding: 8px 16px 16px; gap: 12px; }}
        .next-up {{ margin-top: 12px; padding: 10px 12px; }}
        .next-up-name {{ font-size: 16px; }}
        .stat-number {{ font-size: 22px; }}
        /* Overview block: let its contents shrink/wrap to phone width instead of
           overflowing and getting clipped by .container{{overflow:hidden}}.
           The grid items (overview-left/right) and the inner stat cards need
           min-width:0 so the 1fr column can shrink below their min-content. */
        .overview-left, .overview-right {{ min-width: 0; }}
        .stat-grid > .stat-card {{ min-width: 0; }}
        /* Weather/conditions cards: wrap to multiple rows instead of one no-wrap strip */
        .fc-container {{ flex-wrap: wrap; gap: 8px; }}
        .fc-card {{ flex: 1 1 calc(33.333% - 8px); max-width: none; min-width: 0; }}
        /* Top-15 leaderboard: allow player names to wrap so the table can narrow */
        .top10-table {{ table-layout: fixed; }}
        .top10-table td.player {{ white-space: normal; }}
        /* Season Schedule table: fixed layout + let the date wrap so it fits phone width */
        .schedule-table {{ table-layout: fixed; }}
        .schedule-table td, .schedule-table thead th {{ padding: 5px 4px; }}
        .sched-date {{ white-space: normal; }}
    }}
    /* ── Narrow screens: standings collapses to Rank / +/- / Player / Total ──
       Breakpoint is 1024px (not 768) because the full 12-column table needs
       ~975px to render. Below that it would overflow the viewport, so we show
       the collapsed card with tap-to-expand instead. */
    @media screen and (max-width: 1024px) {{
        /* Hide every standings cell, then re-show the four we keep + the detail row */
        .standings-container table thead th,
        .standings-container table tbody td {{ display: none; }}
        .standings-container table thead th:nth-child(1),
        .standings-container table thead th:nth-child(2),
        .standings-container table thead th:nth-child(3),
        .standings-container table thead th:last-child,
        .standings-container table tbody td:nth-child(1),
        .standings-container table tbody td:nth-child(2),
        .standings-container table tbody td:nth-child(3),
        .standings-container table tbody td:last-child {{ display: table-cell; }}

        .standings-container td.player {{ white-space: normal; }}
        .standings-container tbody tr.has-detail {{ cursor: pointer; }}
        .detail-caret {{
            display: inline-block;
            color: #1a472a;
            font-size: 9px;
            margin-right: 4px;
            vertical-align: middle;
        }}

        .detail-row {{ display: table-row; }}
        .detail-row > td {{
            display: table-cell !important;
            background: #f0f7f2;
            padding: 10px 14px;
        }}
        .detail-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 6px 18px;
        }}
        .detail-item {{
            display: flex;
            justify-content: space-between;
            align-items: baseline;
            font-size: 12px;
            border-bottom: 1px solid #dde7e0;
            padding-bottom: 3px;
        }}
        .detail-label {{ color: #555; }}
        .detail-value {{ font-weight: 700; color: #1a472a; }}
        .detail-value.empty {{ color: #bbb; font-weight: 400; }}

        .mobile-tip {{
            display: block;
            text-align: center;
            font-size: 12px;
            font-weight: 600;
            color: #1a7f3e;
            background: #f0f7f2;
            padding: 8px 12px;
            border-bottom: 1px solid #e0e8e2;
        }}
    }}


    /* ── Smooth anchor scrolling ── */
    html {{ scroll-behavior: smooth; }}
    .container {{ scroll-margin-top: 14px; }}

    /* ── Vertical Table of Contents ── */
    .toc {{
        position: fixed; top: 50%; left: 16px; transform: translateY(-50%);
        z-index: 50; background: #fff; border: 1px solid #dde3da;
        border-radius: 10px; box-shadow: 0 4px 16px rgba(0,0,0,0.10);
        padding: 10px 8px; width: 152px; display: none;
    }}
    .toc-title {{
        font-size: 10px; text-transform: uppercase; letter-spacing: 1px;
        color: #888; font-weight: 700; padding: 2px 10px 8px;
    }}
    .toc a {{
        display: block; padding: 7px 10px; margin: 1px 0;
        font-size: 13px; color: #264636; text-decoration: none;
        border-radius: 6px; border-left: 3px solid transparent;
    }}
    .toc a:hover {{ background: #f0f7f2; }}
    .toc a.active {{
        background: #f0f7f2; color: #1a472a; font-weight: 700;
        border-left-color: #1a472a;
    }}
    .toc-toggle {{
        position: fixed; left: 14px; bottom: 16px; z-index: 51;
        width: 46px; height: 46px; border-radius: 50%;
        background: #1a472a; color: #fff; border: none; cursor: pointer;
        font-size: 20px; box-shadow: 0 3px 12px rgba(0,0,0,0.25);
        align-items: center; justify-content: center; display: none;
    }}
    /* Wide screens: persistent TOC in the left gutter */
    @media (min-width: 1540px) {{
        .toc {{ display: block; }}
    }}
    /* Narrow screens: floating toggle opens the TOC as an overlay */
    @media (max-width: 1539px) {{
        .toc-toggle {{ display: flex; }}
        .toc {{ top: auto; bottom: 72px; left: 14px; transform: none; }}
        .toc.open {{ display: block; }}
    }}

    /* ── Find-your-name search ── */
    .standings-search {{
        display: flex; gap: 8px; align-items: center;
        padding: 12px 16px 6px; max-width: 460px;
    }}
    .standings-search input {{
        flex: 1; padding: 9px 12px; font-size: 14px;
        border: 1px solid #cdd6cf; border-radius: 8px; outline: none;
    }}
    .standings-search input:focus {{ border-color: #1a472a; box-shadow: 0 0 0 2px #e2efe7; }}
    .standings-search button {{
        padding: 9px 12px; border: 1px solid #cdd6cf; border-radius: 8px;
        background: #f5f5f5; cursor: pointer; font-size: 14px; color: #555;
    }}
    .search-count {{ font-size: 12px; color: #888; white-space: nowrap; }}
    tbody.searching tr.has-detail:not(.match) {{ opacity: 0.3; }}
    tr.match > td {{ background: #fff6d6 !important; }}
    tr.match > td.total {{ background: #fceaa8 !important; }}

    /* ── Participation hover checklist ── */
    td.part-cell {{ position: relative; cursor: help; }}
    td.part-cell > span.part-tip {{
        display: none; position: absolute; right: 4px; top: 100%;
        z-index: 60; background: #fff; border: 1px solid #dde3da;
        border-radius: 8px; box-shadow: 0 6px 18px rgba(0,0,0,0.16);
        padding: 8px 10px; min-width: 190px; text-align: left;
    }}
    td.part-cell:hover > span.part-tip {{ display: block; }}
    .part-tip-h {{
        display: block; font-size: 10px; text-transform: uppercase;
        letter-spacing: 0.5px; color: #888; font-weight: 700;
        margin-bottom: 4px; padding-bottom: 4px; border-bottom: 1px solid #eee;
    }}
    .pt-item {{
        display: flex; align-items: center; gap: 6px;
        font-size: 12px; color: #264636; padding: 2px 0; white-space: nowrap;
    }}
    .pt-chk {{ color: #1a7f3e; font-weight: 700; }}
    .pt-pp {{ margin-left: auto; color: #1a472a; font-weight: 700; padding-left: 16px; }}

</style>
</head>
<body>

<nav class="toc" id="toc">
    <div class="toc-title">Jump to</div>
    <a href="#overview" data-sec="overview">Overview</a>
    <a href="#standings" data-sec="standings">Full Standings</a>
    <a href="#schedule" data-sec="schedule">Schedule</a>
    <a href="#spotlight" data-sec="spotlight">Event Spotlight</a>
    <a href="#appendix" data-sec="appendix">Points</a>
</nav>
<button class="toc-toggle" id="tocToggle" aria-label="Jump to section">&#9776;</button>

<!-- PAGE 1: Overview -->
<div class="container" id="overview">
    <div class="header">
        <h1>MGA Ryder Cup Points Standings</h1>
        <div class="subtitle">2025-26 Season - Updated {now}</div>
    </div>
    {build_overview_html(standings, player_data, player_events, rank_changes)}
</div>

<!-- Full Standings -->
<div class="container standings-container" id="standings">
    <div class="header">
        <h1>Full Standings</h1>
        <div class="subtitle">{len(standings)} Players - 2025-26 Season</div>
    </div>
    <div class="mobile-tip">Tap a player to see event-by-event points</div>
    <div class="standings-search">
        <input id="nameSearch" type="text" placeholder="Find your name..." autocomplete="off" spellcheck="false">
        <button id="nameClear" type="button">Clear</button>
        <span class="search-count" id="searchCount"></span>
    </div>
    <div class="table-wrapper">
        <table>
            <thead>
                <tr>
                    <th>Rank</th>
                    <th>+/-</th>
                    <th style="text-align:left; min-width:160px;">Player</th>
                    {tourney_headers}
                    <th>Events</th>
                    <th>Participation Pts</th>
                    <th>Total Pts</th>
                </tr>
            </thead>
            <tfoot>
                <tr>
                    <td colspan="{len(tournament_names) + 6}" class="standings-footer">
                        Participation pts per event: {part_reference}
                    </td>
                </tr>
            </tfoot>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </div>
</div>

<!-- Season Schedule & Points -->
<div class="container schedule-container" id="schedule">
    <div class="header" style="padding: 12px 30px;">
        <h1 style="font-size: 20px;">Season Schedule &amp; Points</h1>
        <div class="subtitle">Full Tournament Calendar</div>
    </div>
    <div class="schedule">
        {build_season_schedule_html()}
    </div>
</div>

<!-- Event Spotlight -->
<div class="container spotlight-container" id="spotlight">
    <div class="header">
        <h1>Event Spotlight: {latest_event_name}</h1>
        <div class="subtitle">Most Recent Tournament</div>
    </div>
    {build_event_spotlight_html(player_data, player_events)}
</div>

<!-- APPENDIX: Points Breakdown -->
<div class="container appendix-container" id="appendix">
    <div class="header">
        <h1>Appendix: Points Breakdown</h1>
        <div class="subtitle">Ryder Cup Points Structure - 2025-26 Season</div>
    </div>
    <div class="appendix">
        {build_appendix_html()}
    </div>
</div>

<script>
(function () {{
    var table = document.querySelector('.standings-container table');
    if (!table) return;

    var heads = table.querySelectorAll('thead th');
    var labels = [];
    for (var i = 0; i < heads.length; i++) labels.push(heads[i].textContent.trim());
    var total = labels.length;
    if (total < 5) return;

    var rows = table.querySelectorAll('tbody tr');
    rows.forEach(function (row) {{
        var cells = row.children;
        if (cells.length < total) return;            // skip footer / non-data rows
        var playerCell = cells[2];
        if (!playerCell) return;

        var caret = document.createElement('span');
        caret.className = 'detail-caret';
        caret.textContent = '\\u25B8';
        playerCell.insertBefore(caret, playerCell.firstChild);
        row.classList.add('has-detail');

        row.addEventListener('click', function () {{
            if (window.innerWidth > 1024) return;    // collapse only active on narrow screens
            var next = row.nextElementSibling;
            if (next && next.classList.contains('detail-row')) {{
                next.parentNode.removeChild(next);
                row.classList.remove('open');
                caret.textContent = '\\u25B8';
                return;
            }}
            var html = '<div class="detail-grid">';
            for (var j = 3; j < total - 1; j++) {{   // columns between Player and Total
                var raw = cells[j] ? cells[j].textContent.trim() : '';
                var isEmpty = (raw === '' || raw === '-');
                html += '<div class="detail-item"><span class="detail-label">' +
                        labels[j] + '</span><span class="detail-value' +
                        (isEmpty ? ' empty' : '') + '">' +
                        (isEmpty ? '\\u2013' : raw) + '</span></div>';
            }}
            html += '</div>';

            var dr = document.createElement('tr');
            dr.className = 'detail-row';
            var td = document.createElement('td');
            td.colSpan = 4;
            td.innerHTML = html;
            dr.appendChild(td);
            row.parentNode.insertBefore(dr, row.nextSibling);
            row.classList.add('open');
            caret.textContent = '\\u25BE';
        }});
    }});
}})();
</script>
<script>
(function () {{
    // ── TOC: active-section highlight + mobile toggle ──
    var toc = document.getElementById('toc');
    var toggle = document.getElementById('tocToggle');
    var links = toc ? Array.prototype.slice.call(toc.querySelectorAll('a')) : [];
    var sections = ['overview','standings','schedule','spotlight','appendix']
        .map(function (id) {{ return document.getElementById(id); }})
        .filter(Boolean);

    if (toggle && toc) {{
        toggle.addEventListener('click', function () {{ toc.classList.toggle('open'); }});
    }}
    links.forEach(function (a) {{
        a.addEventListener('click', function () {{ toc.classList.remove('open'); }});
    }});

    if ('IntersectionObserver' in window && sections.length) {{
        var byId = {{}};
        links.forEach(function (a) {{ byId[a.getAttribute('data-sec')] = a; }});
        var obs = new IntersectionObserver(function (entries) {{
            entries.forEach(function (en) {{
                if (en.isIntersecting) {{
                    links.forEach(function (a) {{ a.classList.remove('active'); }});
                    if (byId[en.target.id]) byId[en.target.id].classList.add('active');
                }}
            }});
        }}, {{ rootMargin: '-45% 0px -50% 0px', threshold: 0 }});
        sections.forEach(function (s) {{ obs.observe(s); }});
    }}

    // ── Find-your-name search ──
    var input = document.getElementById('nameSearch');
    var clearBtn = document.getElementById('nameClear');
    var countEl = document.getElementById('searchCount');
    var tbody = document.querySelector('.standings-container tbody');
    if (input && tbody) {{
        function run() {{
            var q = input.value.trim().toLowerCase();
            var rows = tbody.querySelectorAll('tr.has-detail');
            var first = null, n = 0;
            rows.forEach(function (r) {{ r.classList.remove('match'); }});
            if (!q) {{
                tbody.classList.remove('searching');
                if (countEl) countEl.textContent = '';
                return;
            }}
            tbody.classList.add('searching');
            rows.forEach(function (r) {{
                var pc = r.querySelector('td.player');
                var name = pc ? pc.textContent.toLowerCase() : '';
                if (name.indexOf(q) !== -1) {{
                    r.classList.add('match');
                    n++;
                    if (!first) first = r;
                }}
            }});
            if (countEl) countEl.textContent = n ? (n + ' match' + (n > 1 ? 'es' : '')) : 'no matches';
            if (first) first.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
        }}
        input.addEventListener('input', run);
        if (clearBtn) clearBtn.addEventListener('click', function () {{
            input.value = ''; run(); input.focus();
        }});
    }}
}})();
</script>
<script data-goatcounter="https://mga-eldorado.goatcounter.com/count"
        async src="//gc.zgo.at/count.js"></script>
</body>
</html>"""

    out_path = os.path.join(OUTPUT_DIR, "mga_standings.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML written to {out_path}")

    # Also write index.html for GitHub Pages
    index_path = os.path.join(OUTPUT_DIR, "index.html")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html)

    return out_path


def generate_pdf(html_path):
    """Generate PDF from HTML. Tries Playwright (headless Chromium), then fallbacks."""
    pdf_path = os.path.join(OUTPUT_DIR, "mga_standings.pdf")
    abs_html = os.path.abspath(html_path)

    # Try Playwright (headless Chromium)
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(f"file:///{abs_html.replace(os.sep, '/')}")
            page.pdf(path=pdf_path, format="Letter", landscape=True,
                     margin={"top": "10mm", "bottom": "10mm", "left": "10mm", "right": "10mm"})
            browser.close()
        print(f"PDF written to {pdf_path}")
        return pdf_path
    except Exception:
        pass

    # Try pdfkit / wkhtmltopdf
    try:
        import pdfkit
        pdfkit.from_file(html_path, pdf_path, options={
            "page-size": "Letter",
            "orientation": "Landscape",
            "margin-top": "10mm",
            "margin-bottom": "10mm",
            "margin-left": "10mm",
            "margin-right": "10mm",
        })
        print(f"PDF written to {pdf_path}")
        return pdf_path
    except Exception:
        pass

    # Try weasyprint
    try:
        from weasyprint import HTML
        HTML(filename=html_path).write_pdf(pdf_path)
        print(f"PDF written to {pdf_path}")
        return pdf_path
    except Exception:
        pass

    print("PDF generation skipped - install playwright (py -m pip install playwright && py -m playwright install chromium)")
    print("Or open mga_standings.html in a browser and use Print -> Save as PDF.")
    return None


if __name__ == "__main__":
    player_data, player_events = process_all()
    standings, tournament_names = build_standings(player_data, player_events)
    html_path = generate_html(standings, tournament_names, player_data, player_events)
    generate_pdf(html_path)

    # Quick summary
    print(f"\n{'='*60}")
    print(f"  Total players: {len(standings)}")
    print(f"  Tournaments with data: {sum(1 for t in TOURNAMENTS if t[1] is not None)}")
    if standings:
        print(f"  Leader: {standings[0]['player']} ({standings[0]['total']} pts)")
    print(f"{'='*60}")
